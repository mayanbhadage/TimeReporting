'''
To interact with excel files install openpyxl by
pip install openpyxl
'''
import os
from openpyxl import *
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

class Excel:

    def __init__(self):
        self.wb = None
        self.sheet = None
        self.fileExists = False
        self.isFileCreated = True
        self.path = os.getcwd() + "/MyRecords.xlsx"

    def create_excel_file(self):
        self.wb = Workbook()
        self.sheet = self.wb.get_active_sheet()
        self.sheet.title = "My Time Record"
        self.wb.save("MyRecords.xlsx")


    def open_excel_file(self):
        file_path = os.getcwd()
        self.wb = load_workbook(file_path + "/MyRecords.xlsx")
        self.sheet = self.wb.get_active_sheet()


    def isFileExists(self,path):
        return os.path.isfile(path)

    def create_or_open(self,file):
        if file == False:
            print("File Dosent Exist Creating File... ")
            excel.create_excel_file()
            self.initNewFile()
        else:
            print("Opening File....")
            self.isFileCreated = False
            excel.open_excel_file()


    def initNewFile(self):
        self.sheet['A1'] = "Day"
        self.sheet['B1'] = "Date"
        self.sheet['C1'] = "In Time"
        self.sheet['D1'] = "Out Time"
        self.sheet['E1'] = "Out - In"

        column = 1
        while column < 10:
            i = get_column_letter(column)
            self.sheet.column_dimensions[i].width = 15
            column +=1

        for x in range(1, (self.sheet.max_column) + 1):
            self.sheet.cell(row=1, column=x).alignment = Alignment(horizontal='center', vertical='center',
                                                                  wrap_text=False,shrink_to_fit=True)
        for x in range(1, (self.sheet.max_column)+1):
            self.sheet.cell(row=1, column=x).font = Font(bold=True, size=15,)
        self.wb.save("MyRecords.xlsx")

    def updateDay(self,day=None):
        for cell in self.sheet["A"]:
            if cell.value is None:
                last_row = cell.row
                break
        else:
            last_row = cell.row + 1

        print "Last Empty row is {}".format(last_row)



excel = Excel()

excel.fileExists=excel.isFileExists(excel.path)

excel.create_or_open(excel.fileExists)
excel.updateDay()